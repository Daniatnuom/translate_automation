from google_trans_new import google_translator

def google_translate_e_to_j(sentence):
    translator = google_translator()  
    translate_text = translator.translate(sentence,lang_src='en', lang_tgt='ja')  
    return(translate_text)


# Read sentence from xlxs file at desired colums and put them in the list 

import openpyxl
wb = openpyxl.load_workbook('example_poc.xlsx')
ws = wb.worksheets[0]

trans_en = []
trans_ja = []

# Read Column F and put them in the list
# Function of check null column doesn't implemented yet. 
for cell in ws['F']:
    trans_en.append(cell.value)
#print(trans_en)

# Translate sentence to Japanese and append them in the list name, trans_ja
for i in range(len(trans_en)):
    if i == 0:
        i+1
    else:
        #print(type(trans_en[i]))
        sentence = trans_en[i]
        trans_ja.append(google_translate_e_to_j(sentence))
# # Print translated result 
# for j in range(len(trans_ja)):
#     print(trans_ja[j])

# Now write translated sentence your desired cell in the xlxs file.

wb = openpyxl.load_workbook('example_poc.xlsx')
# Specify the worksheet where you work  
ws = wb.worksheets[0]

for youso in range(len(trans_ja)):
    cell = 'G{}'.format(int(youso)+2)
#     print(cell)
#     print(trans_ja[youso])
    ws[cell] = trans_ja[youso]
#     print(ws[cell])

# save result as your preffered file name. 
# wb.save('your_favorite_filename.xlsx')
wb.save('example_poc_add_jpn.xlsx')